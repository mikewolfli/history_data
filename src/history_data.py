# -*- coding: utf-8 -*-
'''
Created on 2015年8月25日

@author: 10256603
'''
import tkinter as tk
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
from tkinter import scrolledtext
import psycopg2
import ctypes
from ctypes import py_object
from openpyxl import Workbook  
import openpyxl.writer.excel as excel_xlsx
from tkinter import simpledialog

lib=ctypes.cdll.LoadLibrary("./libdeal_res")
lib.combine_list.restype=py_object
lib.combine_new_list.restype=py_object
xls_header=["梯型","计划设计交单期","四位合同号","WBS项目号","合同名称","梯号","项目类型","是否紧急","配置工程师","交付项目文档\n项目开始","项目分配","备注","BOM ransfer和项目release","(预计)发运日期","配置完成刷新ID","梯数","action_id", "flow_ser","状态"]
  
class ScrolledTextDlg(simpledialog.Dialog):
    def __init__(self, title, method=0, parent=None, initialvalue=None):  
        if not parent:
            parent = tk._default_root
        
        self.initialvalue = initialvalue
        self.method=method
        
        simpledialog.Dialog.__init__(self, parent, title)
        
    def body(self, master):
        list_title= Label( master, text='WBS list')
        list_title.pack()
        self.textfield = scrolledtext.ScrolledText( master)
        self.textfield.pack()
        self.textfield.bind_all('<Control-v>', self.copy_ev) 
        self.textfield.bind('<Control-V>', self.copy_ev) 
        self.textfield.bind("<Next>", self.change_line)
        self.textfield.bind("<Alt-L>", self.change_line)
          
        if self.initialvalue is not None:
            self.textfield.delete('1.0',END)
            self.textfield.insert(END, self.initialvalue)
                    
        return self.textfield
    
    def change_line(self, event):
        self.textfield.insert(END,'\n')
        
    def validate(self):
        try:
            result = self.getresult()
        except ValueError:
            messagebox.showwarning(
                "Illegal value",
                self.errormessage + "\nPlease try again",
                parent = self
            )
            return 0
        
        res_list = result.split('\n')
        
        res_res=[]
        count=0
        for res in res_list:
            if len(res.rstrip())==0:
                continue
            
            if len(res.rstrip())!=9 and self.method==0:
                messagebox.showwarning("Illegal value", '物料号字符串长度为9位')
                return 0 
                
            if len(res.rstrip())!=14 and self.method==1:
                messagebox.showwarning("Illegal value", 'WBS No字符串长度为14位')
                return 0                 
            
            if self.method==0:
                l = list(res.rstrip())
                for i in range(len(l) - 1, -1, -1):
                    if not(48 <= ord(l[i]) <= 57): 
                        messagebox.showwarning("Illegal value", '请输入数值')
                        return 0
                    
            count+=1                
            res_res.append(res.rstrip())

        if count==0:
            return 0
                   
        if messagebox.askyesno('是否继续','执行数据数量: '+str(count)+' 条;此操作不可逆，是否继续(YES/NO)?')==NO:
            return 0
        
        self.result=res_res
        return 1
    
    def destroy(self):
        self.textfield=None
        simpledialog.Dialog.destroy(self)        
            
    def getresult(self):
        return self.textfield.get('1.0',END)

    def copy_ev(self, event):
        #self.textfield.delete('1.0',END)
        self.textfield.clipboard_get()
    

def ask_list(title, method=0):
    d=ScrolledTextDlg(title, method)
    return d.result

class Application(Frame):
    def __init__(self, master=None):
        Frame.__init__(self, master)
        self.pack()
        self.createWidgets()
        try:
            self.conn=psycopg2.connect(database="pgworkflow", user="postgres", password="1q2w3e4r", host="10.127.144.62", port="5432")
        except :
            messagebox.showerror(title="连接错误", message="无法连接")
            sys.exit()
        #self.cur = self.conn.cursor()
        
    def quit_func(self):
        if self.conn is not None:
            self.conn.close()
        self.quit()
        
    def save_file(self):
        file_str=filedialog.asksaveasfilename(title="导出文件", initialfile="d:/temp.xlsx",filetypes=[('excel file','.xlsx')])
        if not file_str.endswith(".xlsx"):
            file_str+=".xlsx"
        self.file_string.set(file_str)
    
    def get_old_data(self):
        #cur_his=self.conn.cursor(cursor_factory=extras.DictCursor)
        cur_his=self.conn.cursor()
        
        self.b_new=False
        
        d = ask_list('WBS拷贝器', 1)
        if not d:
            self.b_new=False            
        else:
            self.b_new=True
            str="','".join(d)
            cur_his.execute("select elevator_type,req_configure_finish,req_delivery_date,contract_id,project_id,project_name,lift_no,project_catalog,is_urgent, name,instance_id,action_id,flow_ser,start_date,finish_date,is_active, \
                          to_char((select load from s_unit_parameter where wbs_no=instance_id),'9999') as load,to_char((select speed from s_unit_parameter where wbs_no=instance_id),'9D99') as speed,total_flow,action_name from v_task_out1 WHERE instance_id in ('"+str+"') AND (workflow_id = 'WF0002' OR workflow_id='WF0006') order by instance_id, flow_ser asc;")
            rows_his = cur_his.fetchall()
            #for row in rows_his:
            #    print(row)
            
            self.rows_lib_new2 = lib.combine_new_list(ctypes.py_object(rows_his))
            
        cur_his.execute("select *, to_char((select load from s_unit_parameter where wbs_no=v_history_units_data.wbs_no),'9999') as load,\
                        to_char((select speed from s_unit_parameter where wbs_no=v_history_units_data.wbs_no),'9D99') as speed from v_history_units_data order by wbs_no asc;")
        #col_names = [cn[0] for cn in cur_his.description]
        rows_his = cur_his.fetchall()  

        #print(len(rows_his))
        self.rows_lib= lib.combine_list(ctypes.py_object(rows_his))
        
        cur_his.execute("select elevator_type,req_configure_finish,req_delivery_date,contract_id,project_id,project_name,lift_no,project_catalog, is_urgent, name,instance_id,action_id,flow_ser,start_date,finish_date,is_active, \
                        to_char((select load from s_unit_parameter where wbs_no=instance_id),'9999') as load,to_char((select speed from s_unit_parameter where wbs_no=instance_id),'9D99') as speed,total_flow,action_name from v_task_out1 WHERE status =8 AND (workflow_id = 'WF0002' or workflow_id='WF0006') order by instance_id, flow_ser asc;") 
        rows_his = cur_his.fetchall()       
        rows_temp=lib.combine_new_list(ctypes.py_object(rows_his))
        self.rows_lib.extend(rows_temp)  
        
        cur_his.execute("select elevator_type,req_configure_finish,req_delivery_date,contract_id,project_id,project_name,lift_no,project_catalog,is_urgent, name,instance_id,action_id,flow_ser,start_date,finish_date,is_active, \
                          to_char((select load from s_unit_parameter where wbs_no=instance_id),'9999') as load,to_char((select speed from s_unit_parameter where wbs_no=instance_id),'9D99') as speed,total_flow,action_name from v_task_out1 WHERE status>=1 AND status<8 and status!=4 AND (workflow_id = 'WF0002' OR workflow_id='WF0006') order by instance_id, flow_ser asc;") 
        rows_his = cur_his.fetchall()   
        '''test
        i=0
        s = ''
        for r in rows_his:
            print(r[10])
            if r[10]==s:
                continue
            else:
                s=r[10]
                i+=1
        print(i) 
        ''' 
        self.rows_lib_new=lib.combine_new_list(ctypes.py_object(rows_his))
            
        
    def test_select(self):
        cur_his=self.conn.cursor()
        cur_his.execute("select *, to_char((select load from s_unit_parameter where wbs_no=v_history_units_data.wbs_no),'9999') as load,\
                        to_char((select speed from s_unit_parameter where wbs_no=v_history_units_data.wbs_no),'9D99') as speed from v_history_units_data order by wbs_no asc;")
        #col_names = [cn[0] for cn in cur_his.description]
        rows_his = cur_his.fetchall() 
        
        for row in rows_his:
            print(row)
        
                      
    def export_file(self):
        if self.file_string.get()=="":
            messagebox.showerror("错误","未指定输出文件")
            return
        
        self.get_old_data()
        wb=Workbook()
        ws=wb.worksheets[0]
        ws.title='已发运数据'
        
        for i in range(0,16):
            ws.cell(row=1,column=i+1).value=xls_header[i]
          
        i_row=2  
        for row in self.rows_lib:
            for j in range(0,16):
                ws.cell(row=i_row, column=j+1).value=row[j]
            i_row=i_row+1
              
        ws1=wb.create_sheet()
        ws1.title='未发运数据'
        
        for i in range(0,19):
            ws1.cell(row=1,column=i+1).value=xls_header[i]
            
        i_row=2
        for row in self.rows_lib_new:
            for j in range(0,19):
                ws1.cell(row=i_row,column=j+1).value=row[j]
            i_row=i_row+1

        ws2=wb.create_sheet()
        ws2.title="指定WBS数据"  

        for i in range(0,19):
            ws2.cell(row=1,column=i+1).value=xls_header[i]   
                        
        if self.b_new: 
            i_row=2
            for row in self.rows_lib_new2:
                for j in range(0,19):
                    ws2.cell(row=i_row,column=j+1).value=row[j]
                i_row=i_row+1            
            
        if excel_xlsx.save_workbook(workbook=wb, filename=self.file_string.get()):
            messagebox.showinfo("输出","成功输出!")
                       
    def createWidgets(self):
        self.path_label=Label(self)
        self.path_label["text"]="文件"
        self.path_label.grid(row=0, column=0)
        
        self.file_string=StringVar()
        self.path_entry=Entry(self, textvariable=self.file_string)
        self.path_entry["width"]=50
        self.path_entry.grid(row=0, column=1, columnspan=2)
        self.path_entry["state"]="readonly"
        
        self.path_button=Button(self)
        self.path_button["text"]="变更文件"
        self.path_button.grid(row=0, column=3)
        self.path_button["command"]=self.save_file
        
        self.export_button=Button(self)
        self.export_button["text"]="导出历史数据"
        self.export_button.grid(row=1, column=0)
        self.export_button["command"]=self.export_file
                
        self.quit_button=Button(self)
        self.quit_button["text"]="退出"
        self.quit_button.grid(row=1, column=3)
        self.quit_button["command"]=self.quit_func
        
                
if __name__ == '__main__':
    root=Tk()
    Application(root)
    root.title("历史数据导出程序")
    #root.geometry('640x360')  #设置了主窗口的初始大小960x540  
    root.mainloop()
